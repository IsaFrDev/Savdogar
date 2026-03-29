"""
Bulk operations for products - Excel import/export and bulk updates.
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.utils.text import slugify
from .models import Product


def get_client_ip(request):
    """Get client IP from request."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


def bulk_update_products(products_data, user, ip_address=None):
    """
    Bulk update multiple products at once.
    products_data: list of dicts with 'id' and fields to update (price, stock, etc.)
    Returns: (success_count, errors)
    """
    success_count = 0
    errors = []
    
    for item in products_data:
        product_id = item.get('id')
        if not product_id:
            errors.append({'id': None, 'error': 'Product ID is required'})
            continue
        
        try:
            product = Product.objects.get(id=product_id)
            
            # Update fields
            for field in ['price', 'stock', 'name', 'active', 'compare_price', 'cost_price']:
                if field in item:
                    setattr(product, field, item[field])
            
            product.save()
            success_count += 1
            
        except Product.DoesNotExist:
            errors.append({'id': product_id, 'error': 'Product not found'})
        except Exception as e:
            errors.append({'id': product_id, 'error': str(e)})
    
    return success_count, errors


def export_products_to_excel(products):
    """
    Export products to Excel file.
    Returns BytesIO buffer with Excel file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['ID', 'Name', 'SKU', 'Category', 'Price', 'Compare Price', 'Cost Price', 'Stock', 'Active']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row, product in enumerate(products, 2):
        ws.cell(row=row, column=1, value=product.id).border = thin_border
        ws.cell(row=row, column=2, value=product.name).border = thin_border
        ws.cell(row=row, column=3, value=product.sku).border = thin_border
        ws.cell(row=row, column=4, value=product.category.name if product.category else '').border = thin_border
        ws.cell(row=row, column=5, value=float(product.price)).border = thin_border
        ws.cell(row=row, column=6, value=float(product.compare_price) if product.compare_price else '').border = thin_border
        ws.cell(row=row, column=7, value=float(product.cost_price) if product.cost_price else '').border = thin_border
        ws.cell(row=row, column=8, value=product.stock).border = thin_border
        ws.cell(row=row, column=9, value='Yes' if product.active else 'No').border = thin_border
    
    # Adjust column widths
    column_widths = [8, 40, 15, 20, 12, 15, 15, 10, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def import_products_from_excel(file, store, user, ip_address=None):
    """
    Import products from Excel file.
    Returns: (created_count, updated_count, errors)
    """
    wb = load_workbook(file)
    ws = wb.active
    
    created_count = 0
    updated_count = 0
    errors = []
    
    # Get headers from first row
    headers = [cell.value for cell in ws[1]]
    
    # Map headers to indices
    header_map = {header.lower().strip() if header else '': idx for idx, header in enumerate(headers)}
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        try:
            # Get values from row
            product_id = row[header_map.get('id', 0)] if 'id' in header_map else None
            name = row[header_map.get('name', 1)] if 'name' in header_map else None
            sku = row[header_map.get('sku', 2)] if 'sku' in header_map else ''
            price = row[header_map.get('price', 4)] if 'price' in header_map else 0
            compare_price = row[header_map.get('compare price', 5)] if 'compare price' in header_map else None
            cost_price = row[header_map.get('cost price', 6)] if 'cost price' in header_map else None
            stock = row[header_map.get('stock', 7)] if 'stock' in header_map else 0
            active_str = row[header_map.get('active', 8)] if 'active' in header_map else 'Yes'
            
            if not name:
                continue  # Skip empty rows
            
            active = active_str in ['Yes', 'yes', 'YES', True, 1, '1']
            
            # Update existing or create new
            if product_id:
                try:
                    product = Product.objects.get(id=product_id, store=store)
                    product.name = name
                    product.sku = sku or ''
                    product.price = price
                    product.compare_price = compare_price if compare_price else None
                    product.cost_price = cost_price if cost_price else None
                    product.stock = stock or 0
                    product.active = active
                    product.save()
                    updated_count += 1
                except Product.DoesNotExist:
                    errors.append(f"Row {row_idx}: Product ID {product_id} not found")
            else:
                # Create new product
                Product.objects.create(
                    store=store,
                    name=name,
                    slug=slugify(name),
                    sku=sku or '',
                    price=price or 0,
                    compare_price=compare_price if compare_price else None,
                    cost_price=cost_price if cost_price else None,
                    stock=stock or 0,
                    active=active
                )
                created_count += 1
        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
    
    return created_count, updated_count, errors
